#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Анализатор файлов в текущей папке.
Читает настройки из КаталогФайлов.xlsx, анализирует файлы,
записывает результаты обратно в Excel.
"""

import os
import sys
import time
import hashlib
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Ошибка: Не установлена библиотека openpyxl")
    print("Установите: pip install openpyxl")
    sys.exit(1)


class FileAnalyzer:
    def __init__(self):
        self.excel_file = "КаталогФайлов.xlsx"
        self.start_time = None
        self.max_seconds = 30
        self.in_transaction = True
        self.results = []
        self.total_size_bytes = 0
        self.file_count = 0
        self.errors = []
        
    def read_settings(self):
        """Читает настройки из Excel файла"""
        try:
            wb = load_workbook(self.excel_file)
            ws = wb["Настройки"]
            
            # Проходим по строкам листа Настройки
            for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
                if not row[0]:  # Пропускаем пустые строки
                    continue
                    
                param = str(row[0]).strip()
                value = row[1]
                if value is not None:
                    value = str(value).strip()
                
                if param == "Отсечка секунд":
                    try:
                        self.max_seconds = int(value) if value else 30
                    except ValueError:
                        self.max_seconds = 30
                elif param == "В транзакции":
                    self.in_transaction = value.upper() == "ДА" if value else True
            
            wb.close()
            print(f"Настройки загружены: Отсечка={self.max_seconds}с, Транзакция={'ДА' if self.in_transaction else 'НЕТ'}")
            
        except Exception as e:
            print(f"Ошибка при чтении настроек: {e}")
            print("Использую настройки по умолчанию")
    
    def calculate_md5(self, filepath):
        """Вычисляет MD5 хеш файла"""
        try:
            hash_md5 = hashlib.md5()
            with open(filepath, "rb") as f:
                # Читаем файл блоками для больших файлов
                for chunk in iter(lambda: f.read(4096), b""):
                    hash_md5.update(chunk)
            return hash_md5.hexdigest()
        except Exception as e:
            self.errors.append(f"Ошибка MD5 для {filepath}: {str(e)}")
            return "ОШИБКА"
    
    def analyze_file(self, filepath, rel_path):
        """Анализирует один файл и возвращает данные"""
        try:
            stat = os.stat(filepath)
            
            # Получаем дату создания (в зависимости от ОС)
            try:
                created_timestamp = os.path.getctime(filepath)
            except:
                created_timestamp = stat.st_ctime
            
            created_date = datetime.fromtimestamp(created_timestamp)
            
            # Имя и расширение
            filename = os.path.basename(filepath)
            name_only, extension = os.path.splitext(filename)
            
            # Размер
            size_bytes = stat.st_size
            size_mb = round(size_bytes / (1024 * 1024), 2)
            
            # MD5 хеш
            md5_hash = self.calculate_md5(filepath)
            
            # Форматируем дату
            formatted_date = created_date.strftime("%d.%m.%Y %H:%M")
            
            return {
                'filename': filename,
                'path': rel_path,
                'created_date': formatted_date,
                'size_mb': size_mb,
                'extension': extension.lower(),
                'hash_md5': md5_hash,
                'size_bytes': size_bytes
            }
            
        except Exception as e:
            self.errors.append(f"Ошибка анализа {filepath}: {str(e)}")
            return None
    
    def scan_directory(self):
        """Сканирует текущую директорию"""
        print(f"Начинаю анализ текущей папки: {os.getcwd()}")
        self.start_time = time.time()
        
        try:
            # Получаем все файлы в текущей папке и подпапках
            for root, dirs, files in os.walk('.'):
                # Игнорируем скрытые папки (начинающиеся с .)
                dirs[:] = [d for d in dirs if not d.startswith('.')]
                
                for filename in files:
                    # Проверяем время
                    if time.time() - self.start_time > self.max_seconds:
                        print(f"Достигнут лимит времени ({self.max_seconds}с)")
                        return
                    
                    # Игнорируем скрытые файлы и сам Excel файл
                    if filename.startswith('.') or filename == self.excel_file:
                        continue
                    
                    full_path = os.path.join(root, filename)
                    
                    # Пропускаем папки и специальные файлы
                    if not os.path.isfile(full_path):
                        continue
                    
                    # Относительный путь
                    rel_path = os.path.relpath(full_path, '.')
                    
                    # Анализируем файл
                    file_data = self.analyze_file(full_path, rel_path)
                    
                    if file_data:
                        self.results.append(file_data)
                        self.total_size_bytes += file_data['size_bytes']
                        self.file_count += 1
                        
                        # Выводим прогресс каждые 50 файлов
                        if self.file_count % 50 == 0:
                            print(f"Проанализировано файлов: {self.file_count}")
        
        except KeyboardInterrupt:
            print("\nАнализ прерван пользователем")
        except Exception as e:
            self.errors.append(f"Ошибка сканирования: {str(e)}")
    
    def write_results(self):
        """Записывает результаты в Excel файл"""
        try:
            # Создаем резервную копию если в транзакции
            backup_file = None
            if self.in_transaction:
                backup_file = f"{self.excel_file}.backup"
                import shutil
                shutil.copy2(self.excel_file, backup_file)
                print(f"Создана резервная копия: {backup_file}")
            
            # Загружаем или создаем workbook
            try:
                wb = load_workbook(self.excel_file)
            except:
                wb = Workbook()
            
            # --- Сводная страница ---
            if "Сводная" not in wb.sheetnames:
                ws_summary = wb.create_sheet("Сводная")
                ws_summary.append(["Параметр", "Значение", "Описание"])
                ws_summary.append(["Всего файлов", "", "Количество проанализированных файлов"])
                ws_summary.append(["Общий объем", "", "Суммарный размер всех файлов"])
                ws_summary.append(["Дата анализа", "", "Когда был проведен анализ"])
                ws_summary.append(["Время анализа", "", "Сколько времени было потрачено на анализ"])
                ws_summary.append(["Исходная папка", "", "Путь к анализируемой папке"])
            else:
                ws_summary = wb["Сводная"]
            
            # Обновляем сводные данные
            end_time = time.time()
            analysis_time = round(end_time - self.start_time, 2)
            total_size_mb = round(self.total_size_bytes / (1024 * 1024), 2)
            
            summary_data = {
                "Всего файлов": self.file_count,
                "Общий объем": f"{total_size_mb} МБ",
                "Дата анализа": datetime.now().strftime("%d.%m.%Y %H:%M"),
                "Время анализа": f"{analysis_time} сек",
                "Исходная папка": "."
            }
            
            for i, row in enumerate(ws_summary.iter_rows(min_row=2, max_col=1, values_only=False), start=2):
                param_cell = row[0]
                if param_cell.value in summary_data:
                    ws_summary.cell(row=i, column=2, value=summary_data[param_cell.value])
            
            # --- Страница файлов ---
            if "Файлы" not in wb.sheetnames:
                ws_files = wb.create_sheet("Файлы")
                # Удаляем дефолтный лист если он есть
                if "Sheet" in wb.sheetnames:
                    del wb["Sheet"]
            else:
                ws_files = wb["Файлы"]
            
            # Очищаем старые данные (кроме заголовка)
            ws_files.delete_rows(2, ws_files.max_row)
            
            # Заголовки
            headers = ["Имя файла", "Путь к файлу", "Дата создания", "Размер МБ", "Расширение", "Хеш (MD5)"]
            if ws_files.max_row == 0:
                ws_files.append(headers)
            
            # Записываем данные файлов
            for file_data in self.results:
                ws_files.append([
                    file_data['filename'],
                    file_data['path'],
                    file_data['created_date'],
                    file_data['size_mb'],
                    file_data['extension'],
                    file_data['hash_md5']
                ])
            
            # Автонастройка ширины колонок
            for column in ws_files.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_files.column_dimensions[column_letter].width = adjusted_width
            
            # Сохраняем файл
            wb.save(self.excel_file)
            wb.close()
            
            print(f"Результаты сохранены в {self.excel_file}")
            
            # Удаляем резервную копию если всё успешно
            if backup_file and os.path.exists(backup_file):
                os.remove(backup_file)
                print("Резервная копия удалена (успешное завершение)")
            
            return True
            
        except Exception as e:
            print(f"Ошибка при записи результатов: {e}")
            
            # Восстанавливаем из резервной копии если в транзакции
            if self.in_transaction and backup_file and os.path.exists(backup_file):
                try:
                    import shutil
                    shutil.copy2(backup_file, self.excel_file)
                    print("Выполнен откат к резервной копии (транзакция)")
                except Exception as restore_error:
                    print(f"Ошибка при восстановлении из резервной копии: {restore_error}")
            
            return False
    
    def run(self):
        """Основной метод запуска анализа"""
        print("=" * 50)
        print("Анализатор файлов")
        print("=" * 50)
        
        # Проверяем существование Excel файла
        if not os.path.exists(self.excel_file):
            print(f"Ошибка: Файл {self.excel_file} не найден в текущей папке")
            print(f"Текущая папка: {os.getcwd()}")
            return False
        
        # Читаем настройки
        self.read_settings()
        
        # Анализируем файлы
        print("\nНачинаю анализ файлов...")
        self.scan_directory()
        
        # Выводим статистику
        print(f"\nАнализ завершен:")
        print(f"  Файлов проанализировано: {self.file_count}")
        print(f"  Общий размер: {round(self.total_size_bytes / (1024 * 1024), 2)} МБ")
        print(f"  Ошибок: {len(self.errors)}")
        
        if self.errors:
            print("\nОшибки во время анализа:")
            for error in self.errors[:10]:  # Показываем первые 10 ошибок
                print(f"  - {error}")
            if len(self.errors) > 10:
                print(f"  ... и ещё {len(self.errors) - 10} ошибок")
        
        # Записываем результаты
        print("\nСохраняю результаты в Excel...")
        if self.write_results():
            print("\n✓ Готово!")
            return True
        else:
            print("\n✗ Завершено с ошибками")
            return False


def main():
    """Точка входа"""
    analyzer = FileAnalyzer()
    
    try:
        success = analyzer.run()
        if success:
            # Ждем немного перед закрытием (если запущено двойным кликом)
            input("\nНажмите Enter для выхода...")
        else:
            input("\nНажмите Enter для выхода...")
            
    except KeyboardInterrupt:
        print("\n\nПрограмма прервана пользователем")
    except Exception as e:
        print(f"\nКритическая ошибка: {e}")
        import traceback
        traceback.print_exc()
        input("\nНажмите Enter для выхода...")


if __name__ == "__main__":
    main()
