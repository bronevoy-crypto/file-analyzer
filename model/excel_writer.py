import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

class ExcelWriter:
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "File Analysis"
        self._setup_header()
    
    def _setup_header(self):
        """Настройка заголовков таблицы"""
        headers = [
            "File Name", "Path", "Size (KB)", 
            "Extension", "Created", "Modified",
            "Tags", "Category", "Duplicate Group"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
    
    def add_file_data(self, file_data):
        """Добавление данных о файле в таблицу"""
        row = self.ws.max_row + 1
        
        self.ws.cell(row=row, column=1, value=file_data.get('filename', ''))
        self.ws.cell(row=row, column=2, value=file_data.get('path', ''))
        self.ws.cell(row=row, column=3, value=file_data.get('size_kb', 0))
        self.ws.cell(row=row, column=4, value=file_data.get('extension', ''))
        
        # Форматирование дат
        created = file_data.get('created')
        if created:
            self.ws.cell(row=row, column=5, value=created.strftime('%Y-%m-%d %H:%M:%S'))
        
        modified = file_data.get('modified')
        if modified:
            self.ws.cell(row=row, column=6, value=modified.strftime('%Y-%m-%d %H:%M:%S'))
        
        self.ws.cell(row=row, column=7, value=', '.join(file_data.get('tags', [])))
        self.ws.cell(row=row, column=8, value=file_data.get('category', ''))
        self.ws.cell(row=row, column=9, value=file_data.get('duplicate_group', ''))
        
        # Автоматическая подгонка ширины колонок
        for column in self.ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            self.ws.column_dimensions[column_letter].width = adjusted_width
    
    def save(self, target_directory, analysis_name=None):
        """
        Сохраняет Excel файл рядом с анализируемой папкой
        
        Args:
            target_directory: путь к анализируемой папке
            analysis_name: название анализа (необязательно)
        """
        # Генерируем имя файла с временной меткой
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        if analysis_name:
            filename = f"file_analysis_{analysis_name}_{timestamp}.xlsx"
        else:
            # Используем имя папки для названия файла
            folder_name = os.path.basename(os.path.normpath(target_directory))
            filename = f"file_analysis_{folder_name}_{timestamp}.xlsx"
        
        # Сохраняем в той же директории, где находится анализируемая папка
        # (на уровень выше, если нужно прямо рядом)
        parent_dir = os.path.dirname(target_directory)
        
        # Если анализируется корневая директория, сохраняем в нее
        if parent_dir == '':
            parent_dir = target_directory
        
        filepath = os.path.join(parent_dir, filename)
        
        try:
            self.wb.save(filepath)
            return filepath
        except Exception as e:
            # Если не удалось сохранить в указанную директорию,
            # сохраняем в текущей рабочей директории
            print(f"Warning: Could not save to {parent_dir}: {e}")
            fallback_path = os.path.join(os.getcwd(), filename)
            self.wb.save(fallback_path)
            return fallback_path